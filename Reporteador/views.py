#encoding:utf-8
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.db import connection
from django.template import RequestContext
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.db import connections
from openpyxl import Workbook
from openpyxl.styles import Style, Font, Alignment, Border, Side, borders, colors, PatternFill, fills
from openpyxl.styles.numbers import NumberFormatDescriptor, NumberFormat

import json
import datetime

from ConexosAgropecuarios.models import ContratoFondo, DatosFondo, Moneda, ContratoReaseguro
from Direcciones.models import Direccion, Sepomex

@login_required()
def exportarExcelBordereaux(request,IdContratoFondo,IdTipoMoneda,FechaInicio,FechaFinal):
    #Llamamos la consulta para la creacion del reporte
    cursorBordereaux = connections['siobicx'].cursor()
    sql_string = "SELECT * FROM vreportebordereaux WHERE IdContratoFondo = " + IdContratoFondo + " AND IdTipoMoneda = "+IdTipoMoneda +" AND FechaPago Between '"+FechaInicio+" 00:01' AND '"+ FechaFinal +" 23:59'"
    cursorBordereaux.execute(sql_string)
    reporteBordereaux = cursorBordereaux.fetchall()
    #Se crea el libro y se obtiene la hoja, cambiandole el nombre a bordereaux
    libro = Workbook()
    #hoja = libro.get_active_sheet()
    hoja = libro.active
    hoja.title = "Bordereaux"
    #Configuracion de la hoja
    hoja.page_setup.orientation = hoja.ORIENTATION_LANDSCAPE
    #Titulos de la Hoja
    bordereauxSuscripcion = hoja.cell("B2")
    bordereauxSuscripcion.value = "RAMO DE DAÑOS - BORDEREAUX DE SUSCRIPCIÓN"
    bordereauxSuscripcion.style = Style(font=Font(bold=True))
    #Se fucionan las celdas para poner titulos generales del reporte
    hoja.merge_cells("B4:C4")
    hoja.merge_cells("B5:C5")
    hoja.merge_cells("D4:I4")
    hoja.merge_cells("D5:I5")
    hoja.merge_cells("J4:L4")
    hoja.merge_cells("J5:L5")
    hoja.merge_cells("M4:N4")
    hoja.merge_cells("M5:N5")
    hoja.merge_cells("O4:P4")
    hoja.merge_cells("O5:P5")
    hoja.merge_cells("Q4:R4")
    hoja.merge_cells("Q5:R5")
    hoja.merge_cells("S4:T4")
    hoja.merge_cells("S5:T5")
    hoja.merge_cells("U4:V4")
    hoja.merge_cells("U5:V5")
    
    hoja.merge_cells("H8:I9")
    hoja.merge_cells("J8:K9")
    hoja.merge_cells("X9:Y9")
    hoja.merge_cells("Z9:AA9")
    hoja.merge_cells("AB9:AC9")
    hoja.merge_cells("X8:AC8")
    
    #Se definen los estilos para las celdas de los titulos
    estiloDatosGenerales = Style(font=Font(bold=True),alignment=Alignment(horizontal='center'),border=Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK),
                                                                                               right=Side(border_style=borders.BORDER_DOUBLE,color=colors.BLACK),
                                                                                               top=Side(border_style=borders.BORDER_THIN,color=colors.BLACK),
                                                                                               bottom=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)),fill=PatternFill(fill_type=fills.FILL_SOLID,start_color=colors.YELLOW))
    estiloTitulosCeldas = Style(font = Font (bold = True), alignment = Alignment(horizontal = 'center',vertical = 'center'),border = Border(right=Side(border_style=borders.BORDER_THIN,color=colors.BLACK),
                                                                                                                        left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK),
                                                                                                                        top=Side(border_style=borders.BORDER_THIN,color=colors.BLACK),
                                                                                                                        bottom=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)))
    numeroFondoTitulo = hoja.cell("B4")
    numeroFondoTitulo.value = "NÚMERO DE FONDO"    
    nombreFondoTitulo = hoja.cell("D4")
    nombreFondoTitulo.value = "NOMBRE DEL FONDO"
    numeroContratoTitulo = hoja.cell("J4")
    numeroContratoTitulo.value = "NO CONTRATO"
    estadoTitulo = hoja.cell("M4")
    estadoTitulo.value = "ESTADO"
    tipoContratoTitulo = hoja.cell("O4")
    tipoContratoTitulo.value="TIPO DE CONTRATO"
    monedaTitulo = hoja.cell("Q4")
    monedaTitulo.value = "MONEDA"
    fechaCorteTitulo = hoja.cell("S4")
    fechaCorteTitulo.value = "FECHA DE CORTE"
    ejercicioTitulo = hoja.cell("U4")
    ejercicioTitulo.value = "EJERCICIO"
    numeroFondoTitulo.style = nombreFondoTitulo.style = numeroContratoTitulo.style = estadoTitulo.style = tipoContratoTitulo.style = monedaTitulo.style = fechaCorteTitulo.style = ejercicioTitulo.style = estiloDatosGenerales
    #Se pone los datos generales del Fondo y del Contrato de Reaseguro 
    datosFondo = DatosFondo.objects.all()
    for datoFondo in datosFondo:
        NombreFondo = datoFondo.Persona.RazonSocial
        NumeroFondo = datoFondo.NumeroFondo
        direccionesFondo = Direccion.objects.filter(Persona = datoFondo.Persona)
        for direccionFondo in direccionesFondo:
            estadoFondo = Sepomex.objects.using("catalogos").get(IdSepomex = direccionFondo.IdSepomex)
    contratoFondo = ContratoFondo.objects.get(IdContratoFondo = IdContratoFondo)
    tipoMonedaContrato = Moneda.objects.using("catalogos").get(IdMoneda = contratoFondo.IdMoneda)
    tipoMonedaReporte = Moneda.objects.using("catalogos").get(IdMoneda = IdTipoMoneda)
    tipoContratoReaseguro = ContratoReaseguro.objects.using("catalogos").get(IdContratoReaseguro = contratoFondo.IdContratoReaseguro)
    numeroFondo = hoja.cell("B5")
    numeroFondo.value = NumeroFondo
    numeroFondo.style = Style(alignment=Alignment(horizontal='center'))
    nombreFondo = hoja.cell("D5")
    nombreFondo.value = NombreFondo
    nombreFondo.style = Style(alignment=Alignment(horizontal='center'))
    estadoFondoDireccion = hoja.cell("M5")
    estadoFondoDireccion.value = estadoFondo.DEstado.upper()
    estadoFondoDireccion.style = Style(alignment=Alignment(horizontal='center'))
    numeroContrato = hoja.cell("J5")
    numeroContrato.value = contratoFondo.NumeroContrato
    numeroContrato.style = Style(alignment=Alignment(horizontal='center'))
    tipoContrato = hoja.cell("O5")
    tipoContrato.value = tipoContratoReaseguro.RazonContenido
    tipoContrato.style = Style(alignment=Alignment(horizontal='center'))
    moneda = hoja.cell("Q5")
    moneda.value = tipoMonedaContrato.Nombre
    moneda.style = Style(alignment=Alignment(horizontal='center'))    
    fInicio = datetime.datetime.strptime(FechaInicio,"%Y-%m-%d")
    fFin = datetime.datetime.strptime(FechaFinal,"%Y-%m-%d")
    fechaCorte = hoja.cell("S5")
    fechaCorte.value = fInicio.strftime("%d/%m/%Y") + " a " + fFin.strftime("%d/%m/%Y")
    fechaCorte.style = Style(alignment=Alignment(horizontal='center'))
    ejercicio = hoja.cell("U5")
    ejercicio.value = fInicio.year
    ejercicio.style = Style(alignment=Alignment(horizontal='center'))
   
    #Se obtiene las celdas en la cual se colocaran los titulos de los campos
    rango_celdas = hoja.range("B10:AL10")
    #Se crea una tupla con los nombres de los campos
    nombre_campos = "Constancia","Inciso","Endoso","Tipo de Endoso","Socio Asegurado","Producto","Desde","Hasta","Desde","Hasta","Estado","Municipio","Código Postal","No. Unidades de Producción","Nombre del Bien Asegurado","No. de Bienes Asegurados","Valor de los Bienes En Cobertura Limitada","Suma Asegurada Edificios o Instalaciones","Suma Asegurada de Contenidos","Suma Asegurada Total","Ramo","Riesgo Protegido","% ó al %o","Importe","% ó al %o","Importe","%","Importe","Deducible","Participación a Perdida","Numero de Pago","Fecha de Pago","Forma de Pago","Georeferencia","Tipo de Cubierta","No. de Pisos (Niveles)","Observaciones"
    #Se asigna cada nombre de campo a cada celda del rango            
    for campo in rango_celdas:
        indice = 0
        for celda in campo:
            celda.value = nombre_campos[indice]
            celda.style = estiloTitulosCeldas
            indice += 1
    #Se asignan otros titulos
    vigenciaOriginal = hoja.cell("H8")
    vigenciaOriginal.value = "Vigencia Original"
    vigenciaEndoso = hoja.cell("J8")
    vigenciaEndoso.value = "Vigencia Endoso"
    tarifaImporteCuotas = hoja.cell("X8")
    tarifaImporteCuotas.value = "Tarifa e Importe de Cuotas"
    fondoTitle = hoja.cell("X9")
    fondoTitle.value = "Fondo"
    reaseguroTitle = hoja.cell("Z9")
    reaseguroTitle.value = "Reaseguro"
    totalTitle = hoja.cell("AB9")
    totalTitle.value = "Total"
    hoja.cell("L8").style = Style(border = Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK))) #Estas celdas se pone border left porque cuando se fucionan las celdas el border right no se respeta
    hoja.cell("L9").style = Style(border = Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)))
    hoja.cell("W4").style = Style(border = Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)))
    hoja.cell("AD8").style = Style(border = Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)))
    hoja.cell("AD9").style = Style(border = Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)))
    hoja.cell("AM10").style = Style(border = Border(left=Side(border_style=borders.BORDER_THIN,color=colors.BLACK)))
    vigenciaOriginal.style = vigenciaEndoso.style = tarifaImporteCuotas.style = fondoTitle.style = reaseguroTitle.style = totalTitle.style = estiloTitulosCeldas
    #se obtiene el rango de celdas para insertar los datos
    rowsBordereaux = 10 + len(reporteBordereaux)
    celdas_datos = hoja.range("B11:AL{0}".format(rowsBordereaux))
    for fila in range(len(reporteBordereaux)):
        columna = 0
        for celda in celdas_datos[fila]:
            celda.value = reporteBordereaux[fila][columna]
            columna += 1   
    
    rowsTotales = 13 + fila # Para obtener el final y poner los totales se comienza con 10 por los titules y se deja una fila en blanco
    totales = hoja.cell("B{0}".format(rowsTotales))
    totales.value = "TOTAL " + tipoMonedaReporte.Nombre
    sumaAseguradaTotal = hoja.cell("U{0}".format(rowsTotales))
    sumaAseguradaTotal.value = "=SUM(U11:U{0})".format(rowsTotales-1)
    importeFondoTotal = hoja.cell("Y{0}".format(rowsTotales))
    importeFondoTotal.value = "=SUM(Y11:Y{0})".format(rowsTotales-1)
    importeReaseguroTotal = hoja.cell("AA{0}".format(rowsTotales))
    importeReaseguroTotal.value = "=SUM(AA11:AA{0})".format(rowsTotales-1)
    importeTotal = hoja.cell("AC{0}".format(rowsTotales))
    importeTotal.value = "=SUM(AC11:AC{0})".format(rowsTotales-1)
    #Definimos tamaño de las celdas
    hoja.column_dimensions['B'].width = 10
    hoja.column_dimensions['C'].width = 10
    hoja.column_dimensions['D'].width = 10
    hoja.column_dimensions['E'].width = 20
    hoja.column_dimensions['F'].width = 20
    hoja.column_dimensions['G'].width = 30
    hoja.column_dimensions['H'].width = 20
    hoja.column_dimensions['I'].width = 20
    hoja.column_dimensions['J'].width = 25
    hoja.column_dimensions['K'].width = 25
    hoja.column_dimensions['L'].width = 20
    hoja.column_dimensions['M'].width = 20
    hoja.column_dimensions['N'].width = 20
    hoja.column_dimensions['O'].width = 25
    hoja.column_dimensions['P'].width = 25
    hoja.column_dimensions['Q'].width = 25
    hoja.column_dimensions['R'].width = 25
    hoja.column_dimensions['S'].width = 25
    hoja.column_dimensions['T'].width = 25
    hoja.column_dimensions['U'].width = 25
    hoja.column_dimensions['V'].width = 25
    hoja.column_dimensions['W'].width = 25
    hoja.column_dimensions['X'].width = 10
    hoja.column_dimensions['Y'].width = 20
    hoja.column_dimensions['Z'].width = 10
    hoja.column_dimensions['AA'].width = 25
    hoja.column_dimensions['AB'].width = 10
    hoja.column_dimensions['AC'].width = 25
    hoja.column_dimensions['AD'].width = 20
    hoja.column_dimensions['AE'].width = 25
    hoja.column_dimensions['AF'].width = 20
    hoja.column_dimensions['AG'].width = 20
    hoja.column_dimensions['AH'].width = 20
    hoja.column_dimensions['AI'].width = 25
    hoja.column_dimensions['AJ'].width = 25
    hoja.column_dimensions['AK'].width = 25
    hoja.column_dimensions['AL'].width = 25

    #Creamos el archivo en excel
    response = HttpResponse(mimetype="application/ms-excel")
    nombre_archivo = "bordereaux_"+FechaInicio+"_"+FechaFinal+".xlsx"
    contenido = "attachment;filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    libro.save(response)
    return response

@login_required()
def reporteBordereaux(request): # Genera la vista para mostrar la informacion del reporte del bordereaux
    if request.is_ajax(): #Si el get es por ajax se obtiene los parametros para la generacion del reporte bordereaux
        if request.POST['ReporteBordereaux']:
            ReporteBordereaux = json.loads(request.POST['ReporteBordereaux'])
            
            cursorBordereaux = connections['siobicx'].cursor()
            sql_string = "SELECT * FROM vreportebordereaux WHERE IdContratoFondo = " + ReporteBordereaux['IdContratoFondo'] + " AND IdTipoMoneda = "+ReporteBordereaux['IdTipoMoneda']+" AND FechaPago Between '"+ReporteBordereaux['FechaInicio']+" 00:01' AND '"+ ReporteBordereaux['FechaFinal'] +" 23:59'"
            cursorBordereaux.execute(sql_string)
            reporteBordereaux = cursorBordereaux.fetchall()
            rowsBordereaux = len(reporteBordereaux)
            reporte = list()
            for i in range(rowsBordereaux):
                reporte.append({'NumeroConstancia':reporteBordereaux[i][0],'NumeroEndoso':reporteBordereaux[i][2],'TipoEndoso':reporteBordereaux[i][3],'SocioAsegurado':reporteBordereaux[i][4],
                                'Producto':reporteBordereaux[i][5],'VODesde':reporteBordereaux[i][6].strftime("%d/%m/%Y %H:%M:%S"),'VOHasta':reporteBordereaux[i][7].strftime("%d/%m/%Y %H:%M:%S"),'VEDesde':reporteBordereaux[i][8],'VEHasta':reporteBordereaux[i][9],
                                'Estado':reporteBordereaux[i][10],'Municipio':reporteBordereaux[i][11],'CP':reporteBordereaux[i][12],'NombreDelBien':reporteBordereaux[i][14],
                                'NumeroBienes':reporteBordereaux[i][15],'SumaAseguradaContenidos':str(reporteBordereaux[i][18]),'SumaAseguradaTotal':str(reporteBordereaux[i][19]),'Ramo':reporteBordereaux[i][20],
                                'RiesgoProtegido':reporteBordereaux[i][21],'PorcentajeFondo':str(reporteBordereaux[i][22]),'ImporteFondo':str(reporteBordereaux[i][23]),'PorcentajeReaseguro':str(reporteBordereaux[i][24]),
                                'ImporteReaseguro':str(reporteBordereaux[i][25]),'PorcentajeTotal':str(reporteBordereaux[i][26]),'ImporteTotal':str(reporteBordereaux[i][27]),'FechaPago':reporteBordereaux[i][31].strftime("%d/%m/%Y"),
                                'FormaPago':reporteBordereaux[i][32],'Observaciones':reporteBordereaux[i][36]})
            
            return HttpResponse (json.dumps(reporte), content_type="application/json; charset=utf8")
        
    else: #si no es por ajax se muestra el template html
        #se obtienen los contratos de reaseguro que tiene dado de alta el fondo de aseguramiento
        ContratosRE = ContratoFondo.objects.all()
                  
        return render(request, 'reportebordereaux.html', {'ContratosRE':ContratosRE,'usuario':request.user})